#!/usr/bin/env python3
"""
Bot de Casa de Cambio — Telegram + Excel
Recibe mensajes en texto libre, los parsea con IA y los registra en operaciones.xlsx
"""

import os
import json
import logging
import requests
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─── CONFIG ────────────────────────────────────────────────────────────────────
TELEGRAM_TOKEN  = os.getenv("TELEGRAM_TOKEN", "8794992146:AAG5hZxAE0pIDTF6fxl-It11aZtRM1lEKzg")
ANTHROPIC_KEY   = os.getenv("ANTHROPIC_API_KEY", "TU_ANTHROPIC_KEY_AQUI")
EXCEL_FILE      = Path("/tmp/operaciones.xlsx")
STATE_FILE      = Path(__file__).parent / "state.json"
POLL_INTERVAL   = 2  # segundos entre polls
AUTHORIZED_CHAT = 813807479  # Solo Tomas Salgado

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler("bot_cambio.log"), logging.StreamHandler()]
)
log = logging.getLogger(__name__)

# ─── ESTADO ────────────────────────────────────────────────────────────────────

def crear_excel_si_no_existe():
    if EXCEL_FILE.exists():
        return
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Operaciones"
    AZUL_OSC = "1E3A5F"; AZUL_MED = "2E6DA4"; GRIS_CLARO = "F2F5F9"; BLANCO = "FFFFFF"; GRIS_BORDE = "CFD8DC"
    from openpyxl.styles import Side, Border
    def tb():
        s = Side(style="thin", color=GRIS_BORDE)
        return Border(left=s, right=s, top=s, bottom=s)
    ws.merge_cells("A1:N1"); ws["A1"] = "REGISTRO DE OPERACIONES — CASA DE CAMBIO"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=BLANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    headers = [("A","ID",8),("B","FECHA",13),("C","HORA",8),("D","TIPO",9),("E","DIVISA",9),
               ("F","MONTO",14),("G","TIPO CAMBIO",13),("H","TOTAL ARS",15),("I","CLIENTE",18),
               ("J","COMISION %",12),("K","COMISION ARS",14),("L","GANANCIA ARS",14),("M","SALDO USD",13),("N","OBSERVACIONES",22)]
    for col_letter, label, width in headers:
        c = ws[f"{col_letter}4"]; c.value = label
        c.font = Font(name="Arial", bold=True, size=10, color=BLANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_MED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = tb(); ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[4].height = 28
    for row in range(5, 1005):
        fill_color = GRIS_CLARO if row % 2 == 0 else BLANCO
        for col in range(1, 15):
            c = ws.cell(row=row, column=col); c.border = tb()
            c.font = Font(name="Arial", size=10); c.fill = PatternFill("solid", fgColor=fill_color)
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A5"; ws.auto_filter.ref = "A4:N4"
    ws2 = wb.create_sheet("Resumen"); ws2.merge_cells("A1:F1"); ws2["A1"] = "RESUMEN DIARIO"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color=BLANCO)
    ws2["A1"].fill = PatternFill("solid", fgColor=AZUL_OSC)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center"); ws2.row_dimensions[1].height = 28
    rh = [("FECHA",14),("OP. COMPRA",12),("OP. VENTA",12),("TOTAL ARS MOVIDO",18),("COMISION ARS",15),("GANANCIA ARS",15)]
    for i,(h,w) in enumerate(rh,1):
        c = ws2.cell(row=2,column=i,value=h); c.font = Font(name="Arial",bold=True,size=10,color=BLANCO)
        c.fill = PatternFill("solid",fgColor=AZUL_MED); c.alignment = Alignment(horizontal="center"); c.border = tb()
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws3 = wb.create_sheet("Config"); ws3.merge_cells("A1:B1"); ws3["A1"] = "CONFIGURACION"
    ws3["A1"].font = Font(name="Arial",bold=True,size=12,color=BLANCO)
    ws3["A1"].fill = PatternFill("solid",fgColor=AZUL_OSC); ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.column_dimensions["A"].width = 24; ws3.column_dimensions["B"].width = 22
    for i,(k,v) in enumerate([("Comision compra (%)",1.5),("Comision venta (%)",1.5),("Saldo inicial USD",0),("Empresa","Mi Casa de Cambio")],3):
        ck = ws3.cell(row=i,column=1,value=k); ck.font = Font(name="Arial",bold=True,size=10); ck.border = tb()
        cv = ws3.cell(row=i,column=2,value=v); cv.font = Font(name="Arial",size=10,color="0000FF")
        cv.fill = PatternFill("solid",fgColor="FFFDE7"); cv.border = tb()
    wb.save(EXCEL_FILE)
    print(f"Excel creado en {EXCEL_FILE}")

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text())
    return {"last_update_id": 0, "op_counter": 0, "saldo_usd": 0.0}

def save_state(state: dict):
    STATE_FILE.write_text(json.dumps(state, indent=2))

# ─── TELEGRAM ──────────────────────────────────────────────────────────────────
def tg(method: str, **kwargs) -> dict:
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/{method}"
    r = requests.post(url, json=kwargs, timeout=15)
    return r.json()

def send(chat_id: int, text: str, parse_mode="HTML"):
    tg("sendMessage", chat_id=chat_id, text=text, parse_mode=parse_mode)

def get_updates(offset: int) -> list:
    res = tg("getUpdates", offset=offset, timeout=30, limit=10)
    return res.get("result", [])

# ─── PARSEO CON IA ─────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """Sos un asistente para una casa de cambio argentina.
Analizá el mensaje y extraé la operación. Respondé SOLO con JSON válido, sin texto extra.

Formato de respuesta:
{
  "tipo": "COMPRA" o "VENTA",
  "divisa": "USD" | "EUR" | "BRL" | "GBP" | "UYU" | "CLP" | u otra,
  "monto": número flotante (cantidad de divisa extranjera),
  "tipo_cambio": número flotante (pesos por unidad de divisa),
  "cliente": "nombre del cliente o 'Sin nombre' si no se menciona",
  "observaciones": "cualquier dato extra mencionado o vacío",
  "valido": true o false
}

Ejemplos:
- "vendí 10k usd a 1250" → tipo:VENTA, divisa:USD, monto:10000, tipo_cambio:1250
- "compré 500 euros a Juan a 1380" → tipo:COMPRA, divisa:EUR, monto:500, tipo_cambio:1380, cliente:Juan
- "cambié 200 dólares a 1295 con comisión especial" → tipo:VENTA, divisa:USD, monto:200, tipo_cambio:1295, obs:comisión especial
- "hola" → valido:false

COMPRA = el cliente te trae divisas y vos le das pesos (comprás divisas).
VENTA = el cliente te pide divisas y vos se las vendés (vendés divisas)."""

def parsear_operacion(texto: str) -> dict | None:
    try:
        res = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 300,
                "system": SYSTEM_PROMPT,
                "messages": [{"role": "user", "content": texto}],
            },
            timeout=20,
        )
        data = res.json()
        log.info(f"API response status: {res.status_code}")
        if "content" not in data:
            log.error(f"API error response: {data}")
            return None
        raw = data["content"][0]["text"].strip()
        # Limpiar posibles backticks
        raw = re.sub(r"```json|```", "", raw).strip()
        return json.loads(raw)
    except Exception as e:
        log.error(f"Error parseando operación: {e}")
        return None

# ─── EXCEL ─────────────────────────────────────────────────────────────────────
VERDE_BG = "E8F5E9"
ROJO_BG  = "FFEBEE"
AZUL_BG  = "E3F2FD"

def registrar_en_excel(op: dict, state: dict) -> dict:
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Operaciones"]
    ws3 = wb["Config"]

    # Leer comisiones desde Config
    com_compra = float(ws3.cell(row=3, column=2).value or 1.5)
    com_venta  = float(ws3.cell(row=4, column=2).value or 1.5)
    com_pct    = com_compra if op["tipo"] == "COMPRA" else com_venta

    # Calcular
    monto      = float(op["monto"])
    tc         = float(op["tipo_cambio"])
    total_ars  = round(monto * tc, 2)
    com_ars    = round(total_ars * com_pct / 100, 2)
    ganancia   = com_ars  # simplificado; se puede extender con spread

    # Saldo USD
    if op["tipo"] == "COMPRA":
        state["saldo_usd"] = round(state["saldo_usd"] + monto, 2)
    else:
        state["saldo_usd"] = round(state["saldo_usd"] - monto, 2)

    state["op_counter"] += 1
    op_id = f"OP-{state['op_counter']:04d}"

    ahora = datetime.now()
    fila = [
        op_id,
        ahora.strftime("%d/%m/%Y"),
        ahora.strftime("%H:%M"),
        op["tipo"],
        op["divisa"],
        monto,
        tc,
        total_ars,
        op.get("cliente", "Sin nombre"),
        com_pct,
        com_ars,
        ganancia,
        state["saldo_usd"],
        op.get("observaciones", ""),
    ]

    # Encontrar primera fila vacía (desde row 5)
    next_row = 5
    for r in range(5, 1005):
        if ws.cell(row=r, column=1).value is None:
            next_row = r
            break

    color = VERDE_BG if op["tipo"] == "COMPRA" else ROJO_BG
    for col, val in enumerate(fila, 1):
        c = ws.cell(row=next_row, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col in (4,):  # TIPO con color
            c.fill = PatternFill("solid", fgColor=color)
            c.font = Font(name="Arial", size=10, bold=True,
                          color="27500A" if op["tipo"]=="COMPRA" else "791F1F")

    # Actualizar Resumen
    ws2 = wb["Resumen"]
    fecha_hoy = ahora.strftime("%d/%m/%Y")
    fila_resumen = None
    for r in range(3, 500):
        v = ws2.cell(row=r, column=1).value
        if v == fecha_hoy:
            fila_resumen = r
            break
        if v is None:
            fila_resumen = r
            ws2.cell(row=r, column=1, value=fecha_hoy)
            break

    if fila_resumen:
        if op["tipo"] == "COMPRA":
            prev = ws2.cell(row=fila_resumen, column=2).value or 0
            ws2.cell(row=fila_resumen, column=2, value=prev + 1)
        else:
            prev = ws2.cell(row=fila_resumen, column=3).value or 0
            ws2.cell(row=fila_resumen, column=3, value=prev + 1)
        prev_ars = ws2.cell(row=fila_resumen, column=4).value or 0
        ws2.cell(row=fila_resumen, column=4, value=prev_ars + total_ars)
        prev_com = ws2.cell(row=fila_resumen, column=5).value or 0
        ws2.cell(row=fila_resumen, column=5, value=prev_com + com_ars)
        prev_gan = ws2.cell(row=fila_resumen, column=6).value or 0
        ws2.cell(row=fila_resumen, column=6, value=prev_gan + ganancia)

    wb.save(EXCEL_FILE)

    return {
        "op_id": op_id,
        "total_ars": total_ars,
        "com_ars": com_ars,
        "ganancia": ganancia,
        "saldo_usd": state["saldo_usd"],
        "com_pct": com_pct,
    }

# ─── MENSAJE DE CONFIRMACIÓN ───────────────────────────────────────────────────
def mensaje_confirmacion(op: dict, calc: dict) -> str:
    emoji = "🟢" if op["tipo"] == "COMPRA" else "🔴"
    tipo_txt = "COMPRA de divisas" if op["tipo"] == "COMPRA" else "VENTA de divisas"
    return (
        f"{emoji} <b>{calc['op_id']} — {tipo_txt}</b>\n\n"
        f"💱 <b>{op['divisa']}</b>: {op['monto']:,.2f} @ ${op['tipo_cambio']:,.2f}\n"
        f"💵 Total ARS: <b>${calc['total_ars']:,.2f}</b>\n"
        f"📋 Cliente: {op.get('cliente','Sin nombre')}\n"
        f"💼 Comisión ({calc['com_pct']}%): ${calc['com_ars']:,.2f}\n"
        f"📈 Ganancia: <b>${calc['ganancia']:,.2f}</b>\n"
        f"🏦 Saldo USD caja: <b>{calc['saldo_usd']:,.2f}</b>\n\n"
        f"✅ Registrado en Excel — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

# ─── COMANDOS ──────────────────────────────────────────────────────────────────
def cmd_saldo(chat_id, state):
    send(chat_id,
        f"🏦 <b>Saldo actual USD en caja:</b> {state['saldo_usd']:,.2f}\n"
        f"📊 Operaciones totales: {state['op_counter']}")

def cmd_ayuda(chat_id):
    send(chat_id,
        "🤖 <b>Bot Casa de Cambio</b>\n\n"
        "<b>Registrar operación:</b> escribí en texto libre\n"
        "Ejemplos:\n"
        "• <i>vendí 5000 usd a 1280</i>\n"
        "• <i>compré 200 euros a maria a 1390</i>\n"
        "• <i>cambié 10k dólares a 1265 cliente: empresa XYZ</i>\n\n"
        "<b>Comandos:</b>\n"
        "/saldo — ver saldo USD en caja\n"
        "/ayuda — ver este mensaje\n\n"
        "Cada operación se guarda automáticamente en el Excel 📁")

# ─── LOOP PRINCIPAL ────────────────────────────────────────────────────────────
def main():
    crear_excel_si_no_existe()
    state = load_state()
    log.info("🏦 Bot Casa de Cambio iniciado")

    while True:
        try:
            updates = get_updates(state["last_update_id"] + 1)
        except Exception as e:
            log.warning(f"Error obteniendo updates: {e}")
            import time; import time as t; t.sleep(5); continue

        for upd in updates:
            state["last_update_id"] = upd["update_id"]
            msg = upd.get("message", {})
            chat_id = msg.get("chat", {}).get("id")
            texto = msg.get("text", "").strip()

            if not chat_id or not texto:
                continue

            # Solo responde a Tomas
            if chat_id != AUTHORIZED_CHAT:
                send(chat_id, "⛔ No autorizado.")
                continue

            log.info(f"Mensaje de {chat_id}: {texto}")

            if texto.startswith("/start") or texto.startswith("/ayuda"):
                cmd_ayuda(chat_id)
            elif texto.startswith("/saldo"):
                cmd_saldo(chat_id, state)
            else:
                send(chat_id, "⏳ Procesando operación...")
                op = parsear_operacion(texto)

                if not op or not op.get("valido"):
                    send(chat_id,
                        "❓ No entendí la operación. Probá con:\n"
                        "<i>vendí 1000 usd a 1270</i>\n"
                        "<i>compré 500 euros a Juan a 1385</i>")
                else:
                    try:
                        calc = registrar_en_excel(op, state)
                        save_state(state)
                        send(chat_id, mensaje_confirmacion(op, calc))
                    except Exception as e:
                        log.error(f"Error registrando: {e}")
                        send(chat_id, f"⚠️ Error al guardar en Excel: {e}")

            save_state(state)

        import time
        time.sleep(POLL_INTERVAL)

if __name__ == "__main__":
    main()
